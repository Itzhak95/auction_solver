import math
import pandas
from scipy.optimize import fsolve
import openpyxl as xl
from openpyxl.styles import Alignment


epsilon = 10 ** -11
big_epsilon = 10 ** -9
decimals = 14

# # IMPORT THE INPUTS

# To begin, we ask the user the define the set of possible values:

while True:
    try:
        x = abs(int(input("Enter the maximum possible valuation. ")))
        if x >= 3:
            break
        else:
            print("The maximum valuation must be 3 or higher.")
    except:
        print("The maximum valuation must be an integer.")
values = list(range(x+1))


# By assumption, this is also the set of possible bids:
bids = values


# GETTING THE INPUTS FROM THE USER

uniform = 0


while True:
    uniform = input("Are valuations are uniformly distributed (type either 'yes' or 'no')? ")
    if uniform.lower() == "yes" or uniform.lower() == "no":
        break
    else:
        print("Please enter either 'yes' or 'no' (don't include quotation marks).")

if uniform.lower() == "yes":
    uniform = 1
    f = [1/(x + 1)] * (x + 1)
else:
    while True:
        probabilities = input('Below, please specify the probability that the value is 0, the probability it is 1, and so forth.\n'
                              f'This must be a valid probability distribution and must have {x + 1} elements.\n'
                              'Enter separate each probability with a space but no comma. It may be easiest to copy and paste from Excel.\n').split()
        try:
            for i in list(range(x + 1)):
                f[i] = float(probabilities[i])
            break
        except:
            print(f"You must enter a valid probability distribution with {x + 1} elements (and no commas!)")


# Now we determine the number of bidders.
while True:
    try:
        n = int(input("Please enter the number of bidders. "))
        break
    except:
        print("The number of bidders must be an integer.")


# Finally, we determine the auction structure.

while True:
    try:
        all_pay = int(input("Please enter '0' for the first price auction or '1' for the all-pay auction. "))
        if all_pay == 0 or all_pay == 1:
            break
    except:
        print("You must enter either 0 (first price auction) or 1 (all pay auction). Don't use quotation marks. ")


# EQUILIBRIUM CONSTRUCTION

# To begin, let us generate the CDF of values (denoted F).
i = 0
F = []
while i <= x:
    F.append(sum(f[:i + 1]))
    i += 1

# Now let's specify how the winning chance depends on the jump value


def p_win(j):
    value = math.floor(j)
    if value == 0:
        return 0
    else:
        return (F[value - 1] + (j - value) * f[value]) ** (n - 1)


# Now let's specify the difference in payoffs (k is the previous jump):

if all_pay == 1:
    def payoff_difference(j, k):
        return math.floor(j) * (p_win(j) - p_win(k)) - 1
else:
    def payoff_difference(j, k, b):
        return (math.floor(j) - b) * p_win(j) - (math.floor(j) - b + 1) * p_win(k)


# Now for the main algorithm:


if all_pay == 1:
    def next_jump(k):
        # Let's check if a next jump exists
        if x * (1 - p_win(k)) - 1 <= 0:
            return None
        # Let's evaluate the payoff difference at every integer.
        integer_evaluations = [payoff_difference(value, k) for value in values]
        # Let's check if we can see a solution immediately.
        for element in integer_evaluations:
            if element == 0 and integer_evaluations.index(element) > math.floor(k):
                return integer_evaluations.index(element)
        # Let's find the largest integer v that leads to a negative payoff difference:
        negative_evaluations = [element for element in integer_evaluations if element < 0]
        v = len(negative_evaluations)-1
        # Now we check whether there is a non-integer solution:
        candidate_sol = ((1/v + p_win(k)) ** (1/(n - 1))) / f[v] - (F[v - 1]/f[v]) + v
        if abs(payoff_difference(candidate_sol, k)) <= epsilon:
            return candidate_sol
        else:
            return v + 1
else:
    def next_jump(k, b):
        # Let's check if a next jump exists
        if x - b - (x - b + 1) * p_win(k) <= 0:
            return None
        # Let's evaluate the payoff difference at every integer.
        integer_evaluations = [payoff_difference(value, k, b) for value in values]
        # Let's check if we can see a solution immediately.
        for element in integer_evaluations:
            if abs(element - 0) <= epsilon and integer_evaluations.index(element) > math.floor(k):
                return integer_evaluations.index(element)
        # Let's find the largest integer v that leads to a negative payoff difference:
        negative_evaluations = [integer_evaluations[value] for value in values if
                                value <= math.floor(k) or integer_evaluations[value] < 0]
        v = len(negative_evaluations) - 1

        # Now we check whether there is a non-integer solution:
        def h(j):
            return (v - b) * pow(float(F[v - 1]) + float(f[v]) * (j - v), n - 1) - (v - b + 1) * float(p_win(k))
        candidate_sol = fsolve(h, v)
        if abs(payoff_difference(candidate_sol, k, b)) <= epsilon:
            return candidate_sol
        else:
            return v + 1


# Now we apply the function recursively
if all_pay == 1:
    k = 0
    jumps = []
    while True:
        jumps.append(float(k))
        if next_jump(k) is None:
            break
        else:
            k = next_jump(k)
else:
    k = 2
    b = 2
    jumps = [0]
    while True:
        jumps.append(round(float(k), decimals))
        if next_jump(k, b) is None:
            break
        else:
            k = next_jump(k, b)
            b += 1

print(f'The jump vector is \n {jumps}')

# CONVERTING INTO A BEHAVIOURAL STRATEGY

# Let's compute the v vector

v = [math.floor(jump) for jump in jumps]

# For every value, we now obtain a lower bound on the bids (this number may or may not be bid).


def lower_bound(value):
    max_lower = -1
    for element in v:
        if element < value:
            max_lower += 1
    if value == 0:
        max_lower = 0
    return max_lower


# We now obtain a STRICT upper bound (this number won't be bid; but the bid just below will be).


def upper_bound(value):
    min_higher = 0
    for element in v:
        if element <= value:
            min_higher += 1
    return min_higher


# We thus obtain the "support" for every value (which may include one bid that isn't submitted at the start)

support = list(range(x + 1))

for value in values:
    support[value] = list(range(lower_bound(value), upper_bound(value)))


# Now we get the probabilities

p_vector = []

for value in values:
    if math.floor(value) not in v[1:]:
        p_vector.append(1)
    else:
        cumulative = [jump - math.floor(jump) for jump in jumps if math.floor(jump) == value]
        probabilities = list(range(len(cumulative) + 1))
        probabilities[0] = round(cumulative[0], decimals)
        for i in range(1, len(cumulative)):
            probabilities[i] = round(cumulative[i] - cumulative[i - 1], decimals)
        probabilities[-1] = round(1 - sum(probabilities[:-1]), decimals)
        p_vector.append(probabilities)


# Now we tidy up the support and probability vectors by removing bids submitted with 0 probabilities

for value in values:
    if isinstance(p_vector[value], list):
        if p_vector[value][0] == 0:
            p_vector[value] = p_vector[value][1:]
            support[value] = support[value][1:]


# Compute the 'continuous' predictions

if all_pay == 1:
    def g(value):
        return ((n - 1) / n) * (value ** n / x ** (n - 1))
else:
    def g(value):
        return ((n - 1) / n) * value


if uniform == 1:
    continuous_predictions = [g(value) for value in values]


# Export to Excel

if uniform == 1:
    df = pandas.DataFrame({
        'Value': values,
        'Bids': support,
        'Probabilities': p_vector,
        'Continuous model': continuous_predictions
    })

else:
    df = pandas.DataFrame({
        'Value': values,
        'Bids': support,
        'Probabilities': p_vector
    })


df.to_excel('results.xlsx', index=False)

# Tidy up the results

wb = xl.load_workbook('results.xlsx')
sheet = wb['Sheet1']
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row, 3).value == 1:
        sheet.cell(row, 3).value = "[1.0]"

rows = range(1, sheet.max_row + 1)
columns = range(1, 5)
for row in rows:
    for col in columns:
        sheet.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)


wb.save('results.xlsx')


# CHECKING

# Finally, let's check the equilibrium is correct!

# I will cheat a bit and compute best responses to the jump vector

best_responses = []


def prob_win(b):
    if b == 0:
        return 0
    elif b > len(jumps) - 1:
        return 1
    else:
        return pow(F[math.floor(jumps[b])-1] + (jumps[b] - math.floor(jumps[b])) * f[math.floor(jumps[b])], n - 1)


# Payoff from bidding b with a value v:

if all_pay == 1:
    def pi(v, b):
        return v * prob_win(b) - b
else:
    def pi(v, b):
        return (v - b) * prob_win(b)


# Payoff from bidding b with a value v:
error = 0

for v in values:
    payoffs = [pi(v, b) for b in bids]
    max_payoff = max(payoffs)
    optimal_bids = [b for b in bids if abs(pi(v, b) - max_payoff) <= big_epsilon]
    for bid in support[v]:
        if bid not in optimal_bids:
            error = 1

if error == 0:
    print("Equilibrium checked")
else:
    print("Error -- the equilibrium is incorrect. This may be due to rounding error; if so, you should change the epsilon parameters.")
